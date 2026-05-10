"""
Flask Attendance System — app.py
Manages students, fingerprint enrollment, attendance marking, and student login portal.
"""

import os, random, string, threading, smtplib
from datetime import datetime
from zoneinfo import ZoneInfo
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv

load_dotenv()

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, session, jsonify, send_file)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production-xyz987")

# ─── Timezone ─────────────────────────────────────────────────────────────────
IST = ZoneInfo("Asia/Kolkata")

def now_ist():
    return datetime.now(IST)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{os.path.join(BASE_DIR, 'attendance.db')}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# ─── SMTP ─────────────────────────────────────────────────────────────────────
SMTP_HOST     = os.environ.get("SMTP_HOST",     "smtp.gmail.com")
SMTP_PORT     = int(os.environ.get("SMTP_PORT", 587))
SMTP_USER     = os.environ.get("SMTP_USER",     "sensordatadashboard@gmail.com")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "fqvd wnbe oidh akjy")
SMTP_FROM     = os.environ.get("SMTP_FROM",     SMTP_USER)

db = SQLAlchemy(app)

# ─── ESP32 command queue ───────────────────────────────────────────────────────
esp32_command = {"command": None, "fingerprint_id": None}
esp32_lock    = threading.Lock()

# ─── ESP32 API key ─────────────────────────────────────────────────────────────
ESP32_API_KEY = os.environ.get("ESP32_API_KEY", "your-secret-api-key-here")

def esp32_auth(f):
    """Decorator that validates the X-API-Key header on ESP32-facing endpoints."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if request.headers.get("X-API-Key") != ESP32_API_KEY:
            return jsonify({"error": "unauthorized"}), 401
        return f(*args, **kwargs)
    return wrapper


# ═══════════════════════════════════════════════════════════════════════════════
# Models
# ═══════════════════════════════════════════════════════════════════════════════

class Student(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    name           = db.Column(db.String(120), nullable=False)
    roll_no        = db.Column(db.String(30), unique=True, nullable=False)
    email          = db.Column(db.String(120), nullable=True)
    course         = db.Column(db.String(80), nullable=False)
    branch         = db.Column(db.String(80), nullable=False)
    year           = db.Column(db.String(10), nullable=False)
    fingerprint_id = db.Column(db.Integer, unique=True, nullable=True)
    username       = db.Column(db.String(60), unique=True, nullable=False)
    password_hash  = db.Column(db.String(256), nullable=False)
    enrolled_at    = db.Column(db.DateTime, default=now_ist)
    is_enrolled    = db.Column(db.Boolean, default=False)
    records        = db.relationship("AttendanceRecord", backref="student", lazy=True)

    def set_password(self, raw):
        self.password_hash = generate_password_hash(raw)

    def check_password(self, raw):
        return check_password_hash(self.password_hash, raw)

    def to_dict(self):
        return {
            "id":          self.id,
            "name":        self.name,
            "roll_no":     self.roll_no,
            "email":       self.email or "",
            "course":      self.course,
            "branch":      self.branch,
            "year":        self.year,
            "username":    self.username,
            "is_enrolled": self.is_enrolled,
        }


class AttendanceRecord(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"), nullable=False)
    action     = db.Column(db.String(3), nullable=False)
    timestamp  = db.Column(db.DateTime, default=now_ist)
    date       = db.Column(db.String(10))


class Admin(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(60), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)

    def set_password(self, raw):
        self.password_hash = generate_password_hash(raw)

    def check_password(self, raw):
        return check_password_hash(self.password_hash, raw)


# ═══════════════════════════════════════════════════════════════════════════════
# FIX 1 — DB init at module level so Gunicorn creates tables on import.
# The old code put create_defaults() inside `if __name__ == "__main__"` which
# means Gunicorn (which imports the module, never runs it as __main__) never
# called db.create_all(), so no tables were ever created and every request
# crashed with "no such table: student".
# ═══════════════════════════════════════════════════════════════════════════════
def create_defaults():
    db.create_all()
    if not Admin.query.filter_by(username="admin").first():
        a = Admin(username="admin")
        a.set_password("admin123")
        db.session.add(a)
        db.session.commit()
        print("Default admin created — user: admin, pass: admin123")

# Run inside an app context so this works whether imported by Gunicorn or
# executed directly with `python app.py`.
with app.app_context():
    create_defaults()


# ═══════════════════════════════════════════════════════════════════════════════
# Auth decorators
# ═══════════════════════════════════════════════════════════════════════════════

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return wrapper


def student_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("student_id"):
            return redirect(url_for("student_login"))
        return f(*args, **kwargs)
    return wrapper


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def generate_credentials(name: str, roll_no: str):
    base     = name.lower().replace(" ", ".") + "." + roll_no.lower()
    suffix   = "".join(random.choices(string.digits, k=3))
    username = base + suffix
    password = "".join(random.choices(string.ascii_letters + string.digits, k=10))
    return username, password


def next_fingerprint_slot() -> int:
    used = {s.fingerprint_id for s in Student.query.filter(Student.fingerprint_id.isnot(None)).all()}
    for slot in range(1, 128):
        if slot not in used:
            return slot
    raise RuntimeError("Fingerprint sensor is full (127 max)")


def get_last_action(student_id: int):
    rec = (AttendanceRecord.query
           .filter_by(student_id=student_id)
           .order_by(AttendanceRecord.timestamp.desc())
           .first())
    return rec.action if rec else None


def send_credentials_email(to_email, student_name, username, password):
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "Your Attendance Portal Login Credentials"
        msg["From"]    = SMTP_FROM
        msg["To"]      = to_email

        html = f"""
        <html><body style="margin:0;padding:0;background:#0d0f14;font-family:'Segoe UI',sans-serif">
        <div style="max-width:520px;margin:40px auto;background:#161b24;border:1px solid #252d3d;border-radius:16px;overflow:hidden">
          <div style="background:linear-gradient(135deg,#3b82f6,#06b6d4);padding:32px 36px">
            <h1 style="margin:0;color:#fff;font-size:22px;font-weight:700">🎓 Welcome to FingerAttend</h1>
            <p style="margin:8px 0 0;color:rgba(255,255,255,.8);font-size:14px">Your attendance portal credentials</p>
          </div>
          <div style="padding:32px 36px">
            <p style="color:#e2e8f0;font-size:15px;margin:0 0 24px">Hi <strong>{student_name}</strong>,<br>
            Your account has been created. Use the credentials below to log in.</p>
            <div style="background:#0d0f14;border:1px solid #252d3d;border-radius:10px;padding:20px;margin-bottom:24px">
              <div style="margin-bottom:14px">
                <span style="color:#64748b;font-size:11px;text-transform:uppercase;letter-spacing:.06em">Username</span>
                <div style="font-family:'Courier New',monospace;font-size:16px;color:#3b82f6;margin-top:4px;
                            background:rgba(59,130,246,.1);padding:8px 12px;border-radius:6px;
                            border:1px solid rgba(59,130,246,.3)">{username}</div>
              </div>
              <div>
                <span style="color:#64748b;font-size:11px;text-transform:uppercase;letter-spacing:.06em">Password</span>
                <div style="font-family:'Courier New',monospace;font-size:16px;color:#10b981;margin-top:4px;
                            background:rgba(16,185,129,.1);padding:8px 12px;border-radius:6px;
                            border:1px solid rgba(16,185,129,.3)">{password}</div>
              </div>
            </div>
            <p style="color:#64748b;font-size:13px;margin:0">
              ⚠️ Please keep these credentials safe.
            </p>
          </div>
          <div style="padding:16px 36px;border-top:1px solid #252d3d;text-align:center">
            <p style="color:#475569;font-size:12px;margin:0">FingerAttend — Biometric Attendance System</p>
          </div>
        </div>
        </body></html>
        """
        plain = (f"Hi {student_name},\n\nUsername: {username}\nPassword: {password}\n\n— FingerAttend")
        msg.attach(MIMEText(plain, "plain"))
        msg.attach(MIMEText(html,  "html"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo(); server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM, to_email, msg.as_string())
        return True
    except Exception as e:
        app.logger.error(f"Email send failed: {e}")
        return False


# ═══════════════════════════════════════════════════════════════════════════════
# Admin views
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        admin = Admin.query.filter_by(username=request.form["username"]).first()
        if admin and admin.check_password(request.form["password"]):
            session["admin_logged_in"] = True
            session["admin_name"]      = admin.username
            return redirect(url_for("admin_dashboard"))
        flash("Invalid credentials", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin_dashboard():
    students    = Student.query.order_by(Student.enrolled_at.desc()).all()
    today       = now_ist().strftime("%Y-%m-%d")
    today_count = AttendanceRecord.query.filter_by(date=today, action="IN").count()
    return render_template("admin_dashboard.html",
                           students=students,
                           today_count=today_count,
                           total=len(students))


@app.route("/admin/add-student", methods=["GET", "POST"])
@admin_required
def add_student():
    generated = None
    if request.method == "POST":
        name    = request.form["name"].strip()
        roll_no = request.form["roll_no"].strip()
        email   = request.form.get("email", "").strip()
        course  = request.form["course"].strip()
        branch  = request.form["branch"].strip()
        year    = request.form["year"].strip()

        if Student.query.filter_by(roll_no=roll_no).first():
            flash("Roll number already exists", "error")
            return render_template("add_student.html")

        username, raw_password = generate_credentials(name, roll_no)
        student = Student(name=name, roll_no=roll_no, email=email or None,
                          course=course, branch=branch, year=year, username=username)
        student.set_password(raw_password)
        db.session.add(student)
        db.session.commit()

        email_sent = False
        if email:
            email_sent = send_credentials_email(email, name, username, raw_password)

        generated = {"username": username, "password": raw_password,
                     "student": student, "email_sent": email_sent}
        flash("Student added! Share credentials below.", "success")
    return render_template("add_student.html", generated=generated)


@app.route("/admin/edit-student/<int:student_id>", methods=["POST"])
@admin_required
def edit_student(student_id):
    student  = Student.query.get_or_404(student_id)
    new_roll = request.form["roll_no"].strip()
    existing = Student.query.filter_by(roll_no=new_roll).first()
    if existing and existing.id != student_id:
        flash("Roll number already in use by another student.", "error")
        return redirect(url_for("admin_dashboard"))

    student.name    = request.form["name"].strip()
    student.roll_no = new_roll
    student.email   = request.form.get("email", "").strip() or None
    student.course  = request.form["course"].strip()
    student.branch  = request.form["branch"].strip()
    student.year    = request.form["year"].strip()

    new_password = request.form.get("password", "").strip()
    if new_password:
        student.set_password(new_password)

    db.session.commit()
    flash(f"Student {student.name} updated successfully.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/enroll-fingerprint/<int:student_id>", methods=["POST"])
@admin_required
def enroll_fingerprint(student_id):
    student = Student.query.get_or_404(student_id)
    try:
        fp_slot = next_fingerprint_slot()
    except RuntimeError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    student.fingerprint_id = fp_slot
    db.session.commit()

    with esp32_lock:
        esp32_command["command"]        = "ENROLL"
        esp32_command["fingerprint_id"] = fp_slot

    return jsonify({"ok": True, "slot": fp_slot,
                    "message": f"Slot {fp_slot} reserved. Ask student to place finger."})


@app.route("/admin/resend-credentials/<int:student_id>", methods=["POST"])
@admin_required
def resend_credentials(student_id):
    student = Student.query.get_or_404(student_id)
    if not student.email:
        return jsonify({"ok": False, "error": "No email on file for this student."}), 400
    new_password = "".join(random.choices(string.ascii_letters + string.digits, k=10))
    student.set_password(new_password)
    db.session.commit()
    ok = send_credentials_email(student.email, student.name, student.username, new_password)
    if ok:
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "SMTP error — check server logs."}), 500


@app.route("/admin/attendance")
@admin_required
def attendance_log():
    date_filter = request.args.get("date", now_ist().strftime("%Y-%m-%d"))
    records = (AttendanceRecord.query
               .filter_by(date=date_filter)
               .order_by(AttendanceRecord.timestamp.desc())
               .join(Student).all())
    return render_template("attendance_log.html", records=records, date_filter=date_filter)


@app.route("/admin/export-attendance")
@admin_required
def export_attendance():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    date_filter = request.args.get("date", now_ist().strftime("%Y-%m-%d"))
    records = (AttendanceRecord.query
               .filter_by(date=date_filter)
               .order_by(AttendanceRecord.timestamp.asc())
               .join(Student).all())

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {date_filter}"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    in_fill     = PatternFill("solid", fgColor="D1FAE5")
    out_fill    = PatternFill("solid", fgColor="FEE2E2")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    tc = ws["A1"]
    tc.value     = f"Attendance Report — {date_filter}"
    tc.font      = Font(bold=True, size=13, color="1E3A5F")
    tc.alignment = center
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(["Name", "Roll No", "Course / Branch", "Action", "Time"], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
    ws.row_dimensions[2].height = 20

    for ri, r in enumerate(records, 3):
        fill = in_fill if r.action == "IN" else out_fill
        for ci, val in enumerate([
            r.student.name, r.student.roll_no,
            f"{r.student.course} — {r.student.branch}",
            r.action, r.timestamp.strftime("%H:%M:%S")
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.alignment = center; cell.border = border
        ws.row_dimensions[ri].height = 18

    for col, w in zip("ABCDE", [24, 14, 28, 10, 12]):
        ws.column_dimensions[col].width = w

    sr = len(records) + 3
    ws.cell(row=sr, column=1, value="Total Records").font = Font(bold=True)
    ws.cell(row=sr, column=2, value=len(records))
    ws.cell(row=sr, column=3,
            value=f"IN: {sum(1 for r in records if r.action=='IN')}   "
                  f"OUT: {sum(1 for r in records if r.action=='OUT')}")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"attendance_{date_filter}.xlsx")


@app.route("/admin/delete-student/<int:student_id>", methods=["POST"])
@admin_required
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    AttendanceRecord.query.filter_by(student_id=student_id).delete()
    db.session.delete(student)
    db.session.commit()
    flash(f"Student {student.name} removed.", "info")
    return redirect(url_for("admin_dashboard"))


# ═══════════════════════════════════════════════════════════════════════════════
# Student portal
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/", methods=["GET", "POST"])
def student_login():
    if request.method == "POST":
        student = Student.query.filter_by(username=request.form["username"]).first()
        if student and student.check_password(request.form["password"]):
            session["student_id"]   = student.id
            session["student_name"] = student.name
            return redirect(url_for("student_dashboard"))
        flash("Invalid username or password", "error")
    return render_template("student_login.html")


@app.route("/student/logout")
def student_logout():
    session.clear()
    return redirect(url_for("student_login"))


@app.route("/student/dashboard")
@student_required
def student_dashboard():
    student = Student.query.get(session["student_id"])
    records = (AttendanceRecord.query
               .filter_by(student_id=student.id)
               .order_by(AttendanceRecord.timestamp.desc())
               .limit(50).all())
    days = {}
    for r in records:
        days.setdefault(r.date, []).append(r)
    return render_template("student_dashboard.html", student=student, days=days)


# ═══════════════════════════════════════════════════════════════════════════════
# ESP32 API
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/mark-attendance", methods=["POST"])
@esp32_auth
def api_mark_attendance():
    data       = request.get_json(force=True)
    fp_id      = data.get("fingerprint_id")
    confidence = data.get("confidence", 0)

    student = Student.query.filter_by(fingerprint_id=fp_id, is_enrolled=True).first()
    if not student:
        return jsonify({"error": "not_found"}), 404

    action = "OUT" if get_last_action(student.id) == "IN" else "IN"
    now    = now_ist()
    db.session.add(AttendanceRecord(
        student_id=student.id, action=action,
        timestamp=now, date=now.strftime("%Y-%m-%d")
    ))
    db.session.commit()

    return jsonify({"ok": True, "name": student.name, "roll_no": student.roll_no,
                    "action": action, "time": now.strftime("%H:%M"),
                    "confidence": confidence})


@app.route("/api/esp32/command", methods=["GET"])
@esp32_auth
def api_esp32_command():
    with esp32_lock:
        cmd   = esp32_command["command"]
        fp_id = esp32_command["fingerprint_id"]
        if cmd:
            esp32_command["command"]        = None
            esp32_command["fingerprint_id"] = None
            return jsonify({"command": cmd, "fingerprint_id": fp_id})
    return jsonify({"command": None})


@app.route("/api/esp32/enroll-result", methods=["POST"])
@esp32_auth
def api_enroll_result():
    data    = request.get_json(force=True)
    fp_id   = data.get("fingerprint_id")
    success = data.get("success", False)

    student = Student.query.filter_by(fingerprint_id=fp_id).first()
    if student:
        student.is_enrolled = success
        if not success:
            student.fingerprint_id = None
        db.session.commit()
    return jsonify({"ok": True})


@app.route("/api/enrollment-status/<int:student_id>", methods=["GET"])
@admin_required
def api_enrollment_status(student_id):
    student = Student.query.get_or_404(student_id)
    return jsonify({"is_enrolled": student.is_enrolled})


# ═══════════════════════════════════════════════════════════════════════════════
# Dev server entry point (Gunicorn does NOT use this block)
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)